import csv
import re
import json
from openpyxl import load_workbook
import os
from optparse import OptionParser
import copy

# Parse Options
parser = OptionParser()
parser.add_option("-i", "--input", dest="input", default="../project_data/Final data government finance_041018.xlsx", help="Input file", metavar="FILE")
parser.add_option("-o", "--output", dest="output", default="../output/results.csv", help="Output CSV file", metavar="FILE")
parser.add_option("-j", "--outputjson", dest="outputjson", default="../output/results.json", help="Output json file", metavar="FILE")
parser.add_option("-d", "--dict", dest="dict", default="../output/orgDict.json", help="orgDict JSON file", metavar="FILE")
(options, args) = parser.parse_args()


def float_if_possible(test_input):
    """Function that tries to float a number. IF not possible, return alphanumerics only."""
    if test_input is not None:
        try:
            output = float(test_input)
        except ValueError:
            output = re.sub(r'[^a-zA-Z0-9-_\s]', '', test_input).strip()
        return output
    return None

# Import xlsx data
dir_path = os.path.dirname(os.path.realpath(__file__))
inPath = options.input
try:
    wb = load_workbook(filename=os.path.join(dir_path, inPath), read_only=True, data_only=True)
except:
    raise Exception("Input xlsx path required!")
sheets = wb.sheetnames

# Import Dictionary
try:
    with open(os.path.join(dir_path, options.dict), 'r') as output_file:
        orgDict = json.load(output_file)
except:
    orgDict = {}

# budget reference
budgetDict = {}
budgetDict["Actual"] = "actual"
budgetDict["actual"] = "actual"
budgetDict["Act"] = "actual"
budgetDict["Budget"] = "budget"
budgetDict["Est"] = "actual"
budgetDict["Estimate"] = "actual"
budgetDict["EST"] = "actual"
budgetDict["est"] = "actual"
budgetDict["Estim"] = "actual"
budgetDict["None"] = ""
budgetDict["Prel"] = "actual"
budgetDict["Prelim"] = "actual"
budgetDict["Prel Est"] = "actual"
budgetDict["Prel."] = "actual"
budgetDict["Proj"] = "proj"
budgetDict["proj"] = "proj"
budgetDict["Proj."] = "proj"
budgetDict["Prov"] = "actual"
budgetDict["Revised prog"] = "proj"
budgetDict["Projections"] = "proj"
budgetDict["Projection"] = "proj"
budgetDict["Prog"] = "proj"
budgetDict["Rev"] = "proj"
budgetDict["Staff"] = "proj"
budgetDict["Rev Prog"] = "proj"
budgetDict["Rev Proj"] = "proj"
budgetDict[""] = ""
budgetDict[None] = ""

flatData = []
hierData = {"name": "budget", "children": []}
for sheet in sheets:
    levelDict = {}
    levelDict['l1'] = ""
    levelDict['l2'] = ""
    levelDict['l3'] = ""
    levelDict['l4'] = ""
    levelDict['l5'] = ""
    levelDict['l6'] = ""
    ws = wb[sheet]
    rowIndex = 0
    oldNames = []
    names = []
    levels = []
    years = []
    types = []
    values = []
    country = sheet.strip()
    if country not in orgDict:
        orgDict[country] = {}
    print('Reading sheet: '+country)
    for row in ws.iter_rows():
        if len(row) > 0:
            names.append(row[0].value)
            oldNames.append(row[1].value)
            levels.append(row[2].value)
            colLen = len(row)
            if str(row[1].value).lower() == "year":
                for i in range(3, colLen):
                    val = float_if_possible(row[i].value)
                    if str(val).lower() != 'none':
                        years.append(val)
            if str(row[1].value).lower() == "type":
                for i in range(3, colLen):
                    val = float_if_possible(row[i].value)
                    types.append(val)
            if rowIndex >= 5:
                rowValues = []
                for i in range(3, colLen):
                    val = float_if_possible(row[i].value)
                    rowValues.append(val)
                values.append(rowValues)
            rowIndex += 1
    currency = oldNames[1]
    iso = str(names[0]).strip()
    names = names[5:]
    levels = levels[5:]
    oldNames = oldNames[5:]
    nameLen = len(names)
    yearLen = len(years)

    for i in range(0, nameLen):
        name = str(names[i]).strip()
        oldName = str(oldNames[i]).strip()
        if name in ["", "None", "none"]:
            name = oldName
        level = str(levels[i]).lower().strip()
        if level != 'none':
            for j in range(0, yearLen):
                item = {}
                year = years[j]
                yearType = types[j]
                try:
                    levelDict = orgDict[country][level]
                except KeyError:
                    pass
                level_rank = int(level[1:2])+1
                levelDict['l'+str(level_rank)] = name
                item['l1'] = levelDict['l1']
                item['l2'] = levelDict['l2'] if level_rank >= 2 else ""
                item['l3'] = levelDict['l3'] if level_rank >= 3 else ""
                item['l4'] = levelDict['l4'] if level_rank >= 4 else ""
                item['l5'] = levelDict['l5'] if level_rank >= 5 else ""
                item['l6'] = levelDict['l6'] if level_rank >= 6 else ""

                if level != "l0":
                    item_copy = copy.deepcopy(item)
                    item_copy['l'+str(level_rank)] = ""
                    orgDict[country][level] = item_copy
                item['iso'] = iso
                item['country'] = country
                item['currency'] = currency
                item['year'] = year
                item['type'] = budgetDict[yearType]
                try:
                    item['value'] = values[i][j] if str(values[i][j]).lower() != 'none' else ""
                except IndexError:
                    item['value'] = ""
                if budgetDict[yearType] != "":
                    flatData.append(item)
print('Writing orgDict...')
with open(os.path.join(dir_path, options.dict), 'w') as output_file:
    json.dump(orgDict, output_file, ensure_ascii=False, sort_keys=True, indent=2)
print('Done.')
print('Writing CSV...')
keys = ['country', 'iso', 'year', 'currency', 'type', 'l1', 'l2', 'l3', 'l4', 'l5', 'l6', 'value']
with open(os.path.join(dir_path, options.output), 'w') as output_file:
    dict_writer = csv.DictWriter(output_file, keys)
    dict_writer.writeheader()
    dict_writer.writerows(flatData)
print('Done.')
