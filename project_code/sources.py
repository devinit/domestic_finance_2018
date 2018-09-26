import csv
from openpyxl import load_workbook
import os
from optparse import OptionParser
import pdb

parser = OptionParser()
parser.add_option("-i", "--input", dest="input", default="../project_data/Final data government finance_2609186pm.xlsx", help="Input file", metavar="FILE")
parser.add_option("-o", "--output", dest="output", default="../output/domestic-sources.csv", help="Output CSV file", metavar="FILE")
(options, args) = parser.parse_args()

dir_path = os.path.dirname(os.path.realpath(__file__))
inPath = options.input
try:
    wb = load_workbook(filename=os.path.join(dir_path, inPath))
except:
    raise Exception("Input xlsx path required!")
sheets = wb.sheetnames

sources = []
for sheet in sheets:
    ws = wb[sheet]
    country = sheet.strip()
    iso = ""
    yearCols = {}
    print('Reading sheet: '+country)
    for row in ws.iter_rows():
        colLen = len(row)
        if hasattr(row[0], "column"):
            if row[0].column == "A" and row[0].row == 1:
                iso = row[0].value
            if str(row[1].value).lower() == "year":
                for i in range(3, colLen):
                    if hasattr(row[i], "column"):
                        val = str(row[i].value).strip()
                        col = row[i].column
                        if val.lower() != 'none':
                            yearCols[col] = val
        for cell in row:
            if cell.value:
                if str(cell.value).strip().lower() == "source":
                    try:
                        comment = cell.comment
                    except AttributeError:
                        comment = False
                    if comment:
                        try:
                            year = yearCols[cell.column]
                        except KeyError:
                            year = ""
                        obj = {}
                        obj['id'] = iso
                        obj['year'] = str(year)
                        obj['comment'] = comment.content.replace("\n", " ")
                        obj['pub-date'] = ""
                        try:
                            obj['pub-url'] = obj['comment'][obj['comment'].lower().index("http"):].strip()
                        except ValueError:
                            obj['pub-url'] = ""
                        if "Author:" in obj['comment']:
                            obj['pub-title'] = obj['comment'][8:]
                        else:
                            obj['pub-title'] = obj['comment']
                        if obj['pub-url'] != "":
                            obj['pub-title'] = obj['pub-title'][:obj['pub-title'].lower().index("http")].strip()
                        sources.append(obj)
print('Writing CSV...')
keys = ['id', 'year', 'pub-date', 'pub-title', 'pub-url', 'comment']
with open(os.path.join(dir_path, options.output), 'w') as output_file:
    dict_writer = csv.DictWriter(output_file, keys)
    dict_writer.writeheader()
    dict_writer.writerows(sources)
print('Done.')
